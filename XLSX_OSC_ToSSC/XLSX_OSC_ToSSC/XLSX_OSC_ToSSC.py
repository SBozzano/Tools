from tkinter import filedialog
import numpy as np
import openpyxl
import pandas as pd
import os
import csv

def find (Stringa, Carattere):
    for Indice in range( len(Stringa)):
        if Stringa[Indice:len(Carattere)+Indice] == Carattere:
            return 1
    return -1

def trasp(Matrix):
    Matrix_temp = []
    for i in range(len(Matrix[0])):
        Matrix_temp.append([])

    for k in range(len(Matrix)-1):
        for j in range(len(Matrix[k])):

            Matrix_temp[j].append(Matrix[k][j])
    return Matrix_temp

Text_intro =     ( '<?xml version="1.0"?>\n'
                  '<configuration template="%APPPATH%\\templates\SoftScope.pct" version="3"><data><softScopeSession>\n'
                  '			<commSettings>Modbus:1,1000,J#COM:4,38400,N,8,1,H</commSettings>\n'
                  '			<target>\n'
                  '				<maxTracks>20</maxTracks>\n'
                  '				<maxSamples>1</maxSamples>\n'
                  '				<sampleTimeBase>125000</sampleTimeBase>\n'
                  '				<targetConfig><AxX_402p1 template="AxX\AxX_402p1\AxX_402p1.pct" version="1"/>\n'
                  '				</targetConfig>\n'
                  '			</target>\n'
                  '			<session>\n'
                  '				<acquisitionId>405</acquisitionId>\n'
                  '				<numSamples>585</numSamples>\n'
                  '				<triggerSource>varILoopIdRef</triggerSource>\n'
                  '				<triggerSlope>1</triggerSlope>\n'
                  '				<triggerValue>99</triggerValue>\n'
                  '				<preTriggerTime>320</preTriggerTime>\n'
                  '				<tracks>\n')
TextCentral =    (  '            </tracks>\n'
                   '				<timePrescaler>1</timePrescaler>\n'
                   '				<PLCSymTab/>\n'
                   '			</session>\n'
                   '			<data>\n'
                   '				<graph hscale="2" triggerpos="40"><tracks>\n')
TextEnd = ( '</tracks></graph>\n'
               '			</data>\n'
               '		</softScopeSession>\n'
               '	</data></configuration>\n')
Text00 ='<track><DictObject name="'
Text01 = '" index="0" subindex="0" type="5" um="V" scale="10" offs="0" shortdescr="" descr="" bitNumber=""><origin deviceid="AxX_402p1"/></DictObject></track>\n'

Text10 = '<track name="'
Text11 ='" um="" vscale="1" offset="0" color="'
Text12 ='" note="" timeshift="0">\n<samples>\n'
Text13 = '<sample time="'
Text14 ='">'
Text15 ='</sample>\n'
Text16 = '</samples></track>>\n'

ColorString = ['65535', '16776960', '16711935', '65280', '255', '16711680', '16744576', '33023']
path_to_use = None

#***************************Excel************************************+
# ExcelPath = filedialog.askopenfilename(title="Select file", filetypes=(("Excel files", "*.xlsx"),("Excel files", ".xls"),("OSC files", "*.OSC")))
current_directory = os.getcwd()
ExcelPath = filedialog.askopenfilename(
    title="Select file",
    initialdir=current_directory,
    filetypes=(
        (   "Excel/OSC files", "*.xlsx *.xls *.OSC"),
        ("Excel files", "*.xlsx *.xls"),
        ("OSC files", "*.OSC")
    ))
#ExcelPath = filedialog.askopenfilename(title="Select file", filetypes=(("Text files", "*.txt"),("Text files", ".txt")))


if ExcelPath[-5:] == ".xlsx":

    wb = openpyxl.load_workbook(ExcelPath) # --> da migliorare
    sheet = wb.worksheets[0]
    allCells =np.array([[cell.value for cell in row] for row in sheet.iter_rows()])
    allCells = allCells.T
    path_to_use = ExcelPath

elif ExcelPath[-4:] == ".OSC":
    data = pd.read_csv(filepath_or_buffer=ExcelPath, delimiter='\t')
    data.to_excel(excel_writer=ExcelPath[:-4]+'.xlsx', index=False)
    path_to_use = ExcelPath[:-4]+'.xlsx'

if path_to_use is None:
    pass

else:
    wb = openpyxl.load_workbook(path_to_use)  # --> da migliorare
    sheet = wb.worksheets[0]
    allCells = np.array([[cell.value for cell in row] for row in sheet.iter_rows()])
    allCells = allCells.T

    print("Creating file.SSC...")

    fileSSC = open(path_to_use[:-5] + '.SSC', "w")
    fileSSC.write(Text_intro)

    for name in range(1 , len(allCells)):
        if allCells[name][0] is not None:
            fileSSC.write(Text00 + allCells[name][0] + Text01)

    fileSSC.write(TextCentral)

    for name in range(1 , len(allCells)):
        if allCells[name][0] is not None:
            fileSSC.write(Text10 + allCells[name][0] + Text11 + ColorString[name-1] + Text12)
            for sample in range(1 , len(allCells[0])):
                time_to_use = allCells[0][sample]
                num_str = str(allCells[0][sample])

                # Sostituisci la virgola con un punto
                num_str_corrected = num_str.replace(',', '.')

                # Converti la stringa in un numero float
                time = float(num_str_corrected)
                fileSSC.write( Text13 + str(time) + Text14 + str(allCells[name][sample]) + Text15)

            fileSSC.write(Text16)

    fileSSC.write(TextEnd)

    fileSSC.close()


