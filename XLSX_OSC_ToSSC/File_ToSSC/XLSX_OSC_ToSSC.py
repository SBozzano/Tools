from tkinter import filedialog
import numpy as np
import openpyxl
import pandas as pd
import os
import ssc_texts
import csv
from classes import TracksWindow


# TODO
# popup se le colonne sono piu di 8 con la lista delle tracce selezionabili con una spunta

# popip con scritti i "print", con pulsante "Abort"


def find(string, word_to_find, ignore_font):
    word_to_use = word_to_find
    string_to_use = string
    if ignore_font:
        string_to_use = string.casefold()
        word_to_use = word_to_find.casefold()

    if word_to_use in string_to_use:
        return True
    else:
        return False


path_to_use = None


# ***************************Excel************************************+
# ExcelPath = filedialog.askopenfilename(title="Select file", filetypes=(("Excel files", "*.xlsx"),("Excel files", ".xls"),("OSC files", "*.OSC")))

current_directory = os.getcwd()
parent_directory = os.path.dirname(current_directory)

ExcelPath = filedialog.askopenfilename(
    title="Select file",
    initialdir=current_directory,
    filetypes=(
        ("All files", "*.xlsx *.xls *.OSC *.csv"),
        ("Excel files", "*.xlsx *.xls"),
        ("OSC files", "*.OSC"),
        ("csv files", "*.csv")
    ))
# ExcelPath = filedialog.askopenfilename(title="Select file", filetypes=(("Text files", "*.txt"),("Text files", ".txt")))
tracks_window = None

if ExcelPath[-5:] == ".xlsx":
    # wb = openpyxl.load_workbook(ExcelPath)  # --> da migliorare
    # sheet = wb.worksheets[0]
    # allCells = np.array([[cell.value for cell in row] for row in sheet.iter_rows()])
    # allCells = allCells.T
    #
    # print(allCells)
    path_to_use = ExcelPath

if ExcelPath[-4:] == ".OSC":
    print("Reading file.OSC...")
    data = pd.read_csv(filepath_or_buffer=ExcelPath, delimiter='\t')

    data.to_excel(excel_writer=ExcelPath[:-4] + '.xlsx', index=False)
    path_to_use = ExcelPath[:-4] + '.xlsx'

elif ExcelPath[-4:] == ".csv":
    print("Reading file.csv...")

    data = pd.read_csv(filepath_or_buffer=ExcelPath, delimiter=',')

    data.to_excel(excel_writer=ExcelPath[:-4] + '.xlsx', index=False)
    path_to_use = ExcelPath[:-4] + '.xlsx'

if path_to_use is None:
    pass

else:
    print("Creating file.SSC...")
    wb = openpyxl.load_workbook(path_to_use)  # --> da migliorare
    print("A")
    sheet = wb.worksheets[0]
    print("B")
    allCells = np.array([[cell.value for cell in row] for row in sheet.iter_rows()])
    print("C")
    allCells = allCells.T

    if find(allCells[0][0], "Time", True):
        time_found = True

    else:
        time_found = False

    fileSSC = open(path_to_use[:-5] + '.SSC', "w")
    fileSSC.write(ssc_texts.Text_intro)

    if len(allCells) > 8:
        temp_list = []
        for name in range(int(time_found), len(allCells)):
            temp_list.append(allCells[name][0])

        tracks_window = TracksWindow(temp_list)
        print("taìracce: ", tracks_window.get_names_index())

    if not time_found:
        fileSSC.write(ssc_texts.Text00 + "Time" + ssc_texts.Text01)

    for index in tracks_window.get_names_index():

   # for name in range(int(time_found), len(allCells)):

        if allCells[index][0] is not None:
            fileSSC.write(ssc_texts.Text00 + allCells[index][0] + ssc_texts.Text01)
        else:
            break

    fileSSC.write(ssc_texts.TextCentral)

    for color, index in enumerate(tracks_window.get_names_index(), start=0):
   # for name in range(int(time_found), len(allCells)):

        if allCells[index][0] is not None:
            fileSSC.write(ssc_texts.Text10 + allCells[index][0] + ssc_texts.Text11 + ssc_texts.ColorString[
                color] + ssc_texts.Text12)
            time = 0
            for sample in range(1, len(allCells[0])):
                if time_found:
                    time_to_use = allCells[0][sample]
                    num_str = str(allCells[0][sample])

                    # Sostituisci la virgola con un punto
                    num_str_corrected = num_str.replace(',', '.')

                    # Converti la stringa in un numero float
                    time = float(num_str_corrected)
                else:
                    time += 1
                fileSSC.write(
                    ssc_texts.Text13 + str(time) + ssc_texts.Text14 + str(allCells[index][sample]) + ssc_texts.Text15)

            fileSSC.write(ssc_texts.Text16)
        else:
            break
    fileSSC.write(ssc_texts.TextEnd)

    fileSSC.close()
